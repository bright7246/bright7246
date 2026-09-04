import streamlit as st
import streamlit.components.v1 as components
import pandas as pd
import pdfplumber
import re
from collections import defaultdict
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io

st.set_page_config(
    page_title="IRON WARRANTY", 
    page_icon="🚗",
    layout="wide",
    initial_sidebar_state="collapsed"
)

st.markdown(
    """
    <head>
      <meta property="og:title" content="IRON WARRANTY">
      <meta property="og:description" content="아이언모터스 보증팀 지원 프로그램">
      <meta property="og:image" content="https://dummyimage.com/1200x630/0ea5e9/ffffff.png&text=IRON+WARRANTY">
      <meta property="og:image:width" content="1200">
      <meta property="og:image:height" content="630">
      <meta property="og:type" content="website">
      <meta name="google" content="notranslate">
    </head>
    <div translate="no"></div>
    """,
    unsafe_allow_html=True
)

st.markdown(
    """
    <style>
    div[data-testid="stHorizontalBlock"] div.stButton > button {
        height: 65px !important;
        border-radius: 10px !important;
    }
    div[data-testid="stHorizontalBlock"] div.stButton > button p {
        font-size: 18px !important;
        font-weight: 700 !important;
        line-height: 1.3 !important;
    }
    div.share-btn-wrap div.stButton > button {
        height: 38px !important;
        min-height: 38px !important;
        padding: 4px 12px !important;
        border-radius: 8px !important;
        margin-top: 6px !important;
    }
    div.share-btn-wrap div.stButton > button p {
        font-size: 13px !important;
        font-weight: 600 !important;
        line-height: 1.2 !important;
    }
    button[kind="primary"], div.stDownloadButton > button {
        background-color: #0ea5e9 !important;
        border-color: #0ea5e9 !important;
        color: white !important;
    }
    button[kind="primary"]:hover, div.stDownloadButton > button:hover {
        background-color: #0284c7 !important;
        border-color: #0284c7 !important;
        color: white !important;
    }
    button[kind="primary"]:active, button[kind="primary"]:focus,
    div.stDownloadButton > button:active, div.stDownloadButton > button:focus {
        background-color: #0369a1 !important;
        border-color: #0369a1 !important;
        box-shadow: 0 0 0 0.2rem rgba(14, 165, 233, 0.4) !important;
    }
    </style>
    """,
    unsafe_allow_html=True
)

APP_URL = "https://bright7246-cg4cltxcy2z2ksgwbsod2p.streamlit.app"

@st.dialog("📱 프로그램 공유하기")
def share_modal():
    st.write("스마트폰 카메라로 아래 QR 코드를 비추면 즉시 접속할 수 있습니다.")
    qr_url = f"https://api.qrserver.com/v1/create-qr-code/?size=250x250&data={APP_URL}"
    col_img1, col_img2, col_img3 = st.columns([1, 2, 1])
    with col_img2:
        st.image(qr_url, caption="접속용 QR 코드", use_container_width=True)
    
    st.text_input("프로그램 접속 주소", value=APP_URL, disabled=True)
    
    copy_btn_html = f"""
    <div style="display: flex; justify-content: center; margin-top: 10px;">
        <button id="copy-btn" onclick="copyAppUrl()" style="
            background-color: #0ea5e9;
            color: white;
            border: none;
            padding: 10px 24px;
            font-size: 15px;
            font-weight: bold;
            border-radius: 8px;
            cursor: pointer;
            width: 100%;
            transition: 0.2s;
        ">📋 주소 복사하기</button>
    </div>
    <script>
    function copyAppUrl() {{
        navigator.clipboard.writeText('{APP_URL}').then(function() {{
            const btn = document.getElementById('copy-btn');
            btn.innerText = '✅ 복사 완료!';
            btn.style.backgroundColor = '#10b981';
            setTimeout(function() {{
                btn.innerText = '📋 주소 복사하기';
                btn.style.backgroundColor = '#0ea5e9';
            }}, 2000);
        }}).catch(function(err) {{
            alert('복사에 실패했습니다. 주소를 직접 드래그하여 복사해 주세요.');
        }});
    }}
    </script>
    """
    st.components.v1.html(copy_btn_html, height=65)

head_col1, head_col2 = st.columns([8.5, 1.5])
with head_col1:
    st.title("📊 아이언모터스 보증팀 지원 프로그램")
with head_col2:
    st.markdown('<div class="share-btn-wrap">', unsafe_allow_html=True)
    if st.button("🔗 공유 / QR", use_container_width=True):
        share_modal()
    st.markdown('</div>', unsafe_allow_html=True)

if "current_mode" not in st.session_state:
    st.session_state.current_mode = "MW 보증 비교"

nav_col1, nav_col2, nav_col3 = st.columns(3)

with nav_col1:
    btn_mw = st.button(
        "📋 MW 보증 비교 (PDF vs 엑셀)", 
        use_container_width=True, 
        type="primary" if st.session_state.current_mode == "MW 보증 비교" else "secondary"
    )
    if btn_mw:
        st.session_state.current_mode = "MW 보증 비교"
        st.rerun()

with nav_col2:
    btn_coupon = st.button(
        "🚗 쿠폰 보증 비교 (엑셀 vs 엑셀)", 
        use_container_width=True, 
        type="primary" if st.session_state.current_mode == "쿠폰 보증 비교" else "secondary"
    )
    if btn_coupon:
        st.session_state.current_mode = "쿠폰 보증 비교"
        st.rerun()

with nav_col3:
    btn_labor = st.button(
        "🔧 공임코드 비교 (중복 작업 검증)", 
        use_container_width=True, 
        type="primary" if st.session_state.current_mode == "공임코드 비교" else "secondary"
    )
    if btn_labor:
        st.session_state.current_mode = "공임코드 비교"
        st.rerun()

st.divider()

mode = st.session_state.current_mode

# ────────────────────────────────────────────────────────
# 🛠️ [공통 함수]
# ────────────────────────────────────────────────────────
def read_excel_smart_header(uploaded_file):
    uploaded_file.seek(0)
    df_raw = pd.read_excel(uploaded_file, header=None)
    header_row_idx = 0
    for idx, row in df_raw.iterrows():
        row_str = " ".join(row.dropna().astype(str)).upper()
        if 'CLAIM' in row_str or '차량' in row_str or '공임' in row_str:
            header_row_idx = idx
            break
    uploaded_file.seek(0)
    df = pd.read_excel(uploaded_file, header=header_row_idx)
    return df

def find_col_smart(df, keywords, fallback_idx=None):
    for kw in keywords:
        kw_clean = str(kw).replace(" ", "").upper()
        for col in df.columns:
            col_clean = str(col).replace(" ", "").upper()
            if kw_clean in col_clean:
                return col
    if fallback_idx is not None and fallback_idx < len(df.columns):
        return df.columns[fallback_idx]
    return None

def round_half_up(value):
    return int(value + 0.5)

# ────────────────────────────────────────────────────────
# 📊 [컴포넌트 렌더링] 주문번호와 PDF 금액만 복사 허용
# ────────────────────────────────────────────────────────
def render_side_by_side_tables(df_main, df_diff=None, diff_title="🚨 차액 리스트 (100원 이상)"):
    main_headers = ["No."] + list(df_main.columns)
    
    main_tbody = []
    for idx, row in df_main.iterrows():
        is_total = "총합계" in str(row.iloc[0])
        tr_class = ' class="total-row"' if is_total else ''
        main_tbody.append(f'<tr{tr_class}>')
        main_tbody.append(f'<td class="col-no">{idx}</td>')
        
        for c_idx, val in enumerate(row):
            val_str = str(val)
            if c_idx in [0, 1] and not is_total and val_str != "-":
                align_class = "col-id copyable" if c_idx == 0 else "col-amt copyable"
                main_tbody.append(f'<td class="{align_class}" onclick="copyCell(this)">{val_str}</td>')
            else:
                align_class = "col-id" if c_idx == 0 else ("col-diff" if c_idx == len(row)-1 else "col-amt")
                main_tbody.append(f'<td class="{align_class}">{val_str}</td>')
        main_tbody.append('</tr>')

    diff_section = ""
    if df_diff is not None:
        diff_headers = ["No."] + list(df_diff.columns)
        if len(df_diff) > 0:
            diff_tbody = []
            for idx, row in df_diff.iterrows():
                is_total = "총합계" in str(row.iloc[0])
                tr_class = ' class="total-row"' if is_total else ''
                diff_tbody.append(f'<tr{tr_class}>')
                diff_tbody.append(f'<td class="col-no">{idx}</td>')
                
                if not is_total:
                    diff_tbody.append(f'<td class="col-id copyable" onclick="copyCell(this)">{row.iloc[0]}</td>')
                else:
                    diff_tbody.append(f'<td class="col-id">{row.iloc[0]}</td>')
                    
                diff_tbody.append(f'<td class="col-type">{row.iloc[1]}</td>')
                diff_tbody.append(f'<td class="col-desc">{row.iloc[2]}</td>')
                diff_color = "" if is_total else " diff-red"
                diff_tbody.append(f'<td class="col-diff{diff_color}">{row.iloc[3]}</td>')
                diff_tbody.append('</tr>')
            
            diff_section = (
                '<div class="table-card">'
                f'<div class="card-title">{diff_title}</div>'
                '<div class="scroll-wrap">'
                '<table class="compact-table">'
                '<thead><tr>'
                f'<th class="col-no">{diff_headers[0]}</th>'
                f'<th class="col-id">{diff_headers[1]}</th>'
                f'<th class="col-type">{diff_headers[2]}</th>'
                f'<th class="col-desc">{diff_headers[3]}</th>'
                f'<th class="col-diff">{diff_headers[4]}</th>'
                '</tr></thead>'
                f'<tbody>{"".join(diff_tbody)}</tbody>'
                '</table></div></div>'
            )
        else:
            diff_section = (
                '<div class="table-card">'
                f'<div class="card-title">{diff_title}</div>'
                '<div style="padding: 16px; color: #10b981; font-weight: bold; background: #0f172a; border-radius: 6px; border: 1px solid #334155;">'
                '✅ 차액 100원 이상 발생 항목이 없습니다.'
                '</div></div>'
            )

    css_code = """
      * { box-sizing: border-box; margin: 0; padding: 0; font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, sans-serif; }
      body { background-color: transparent; color: #f8fafc; overflow-x: hidden; }
      .flex-container { display: flex; gap: 20px; align-items: flex-start; justify-content: flex-start; }
      .table-card { flex: 0 0 auto; }
      .card-title { font-size: 18px; font-weight: bold; margin-bottom: 10px; color: #f1f5f9; }
      .scroll-wrap { max-height: 640px; overflow-y: auto; border: 1px solid #334155; border-radius: 6px; }
      .scroll-wrap::-webkit-scrollbar { width: 8px; height: 8px; }
      .scroll-wrap::-webkit-scrollbar-track { background: #0f172a; }
      .scroll-wrap::-webkit-scrollbar-thumb { background: #334155; border-radius: 4px; }
      .compact-table { border-collapse: collapse; width: max-content; font-size: 16px; user-select: text; }
      .compact-table thead th { position: sticky; top: 0; background-color: #1e293b; color: #ffffff; padding: 10px 14px; font-weight: 700; border-bottom: 2px solid #475569; border-right: 1px solid #334155; white-space: nowrap; z-index: 2; }
      .compact-table tbody td { padding: 9px 14px; border-bottom: 1px solid #334155; border-right: 1px solid #334155; white-space: nowrap; transition: background-color 0.15s ease; }
      .compact-table tbody td.copyable { cursor: pointer; }
      .compact-table tbody td.copyable:hover { background-color: rgba(14, 165, 233, 0.25) !important; }
      .total-row { background-color: #0f172a !important; font-weight: bold; color: #38bdf8 !important; }
      .diff-red { color: #ef4444 !important; font-weight: bold; }
      .col-no { min-width: 48px; text-align: center; font-weight: bold; }
      .col-id { min-width: 130px; text-align: center; }
      .col-amt { min-width: 140px; text-align: right; }
      .col-type { min-width: 120px; text-align: center; color: #38bdf8; }
      .col-desc { min-width: 220px; text-align: left; }
      .col-diff { min-width: 110px; text-align: right; }
      #toast { visibility: hidden; position: fixed; top: 14px; left: 50%; transform: translateX(-50%); background-color: #0284c7; color: #ffffff; padding: 9px 18px; border-radius: 6px; font-weight: bold; font-size: 14px; z-index: 999999; box-shadow: 0 4px 12px rgba(0,0,0,0.5); }
      #toast.show { visibility: visible; animation: fadein 0.2s, fadeout 0.3s 1.1s; }
      @keyframes fadein { from { opacity: 0; top: 0px; } to { opacity: 1; top: 14px; } }
      @keyframes fadeout { from { opacity: 1; top: 14px; } to { opacity: 0; top: 0px; } }
    """

    js_code = """
      function copyCell(el) {
        const text = el.innerText.trim();
        if (!text || text === '-') return;
        const ta = document.createElement('textarea');
        ta.value = text;
        ta.style.position = 'fixed';
        ta.style.opacity = '0';
        document.body.appendChild(ta);
        ta.select();
        try { document.execCommand('copy'); } catch(e) { navigator.clipboard.writeText(text); }
        document.body.removeChild(ta);

        const origBg = el.style.backgroundColor;
        el.style.backgroundColor = '#0ea5e9';
        setTimeout(function() { el.style.backgroundColor = origBg; }, 200);

        const toast = document.getElementById('toast');
        toast.innerText = '📋 복사 완료: ' + text;
        toast.className = 'show';
        setTimeout(function() { toast.className = ''; }, 1400);
      }
    """

    full_html = (
        '<!DOCTYPE html><html><head><meta charset="utf-8" />'
        f'<style>{css_code}</style></head>'
        '<body>'
        '<div id="toast">📋 복사 완료!</div>'
        '<div class="flex-container">'
        '<div class="table-card">'
        '<div class="card-title">📋 상세 대조 내역</div>'
        '<div class="scroll-wrap">'
        '<table class="compact-table">'
        '<thead><tr>'
        f'<th class="col-no">{main_headers[0]}</th>'
        f'<th class="col-id">{main_headers[1]}</th>'
        f'<th class="col-amt">{main_headers[2]}</th>'
        f'<th class="col-amt">{main_headers[3]}</th>'
        f'<th class="col-diff">{main_headers[4]}</th>'
        '</tr></thead>'
        f'<tbody>{"".join(main_tbody)}</tbody>'
        '</table></div></div>'
        f'{diff_section}'
        '</div>'
        f'<script>{js_code}</script>'
        '</body></html>'
    )
    
    calc_height = min(680, max(260, len(df_main) * 44 + 90))
    components.html(full_html, height=calc_height, scrolling=False)

# ────────────────────────────────────────────────────────
# 1️⃣ [모드 1] MW 보증 비교
# ────────────────────────────────────────────────────────
def load_excel_mw(uploaded_file):
    df = read_excel_smart_header(uploaded_file)
    col_claim_no = find_col_smart(df, ['CLAIM NO', 'CLAIM_NO', '클레임번호', '청구번호', 'CLAIM'], fallback_idx=0)
            
    target_cols = ['공임청구액', '공임청구부가세', '부품청구액', '부품청구부가세']
    for col in target_cols:
        matched_col = find_col_smart(df, [col])
        if matched_col:
            df[matched_col] = pd.to_numeric(df[matched_col], errors='coerce').fillna(0)
            
    c_labor = find_col_smart(df, ['공임청구액'])
    c_labor_vat = find_col_smart(df, ['공임청구부가세'])
    c_part = find_col_smart(df, ['부품청구액'])
    c_part_vat = find_col_smart(df, ['부품청구부가세'])

    df['Excel_Total'] = (
        (df[c_labor] if c_labor else 0) + 
        (df[c_labor_vat] if c_labor_vat else 0) + 
        (df[c_part] if c_part else 0) + 
        (df[c_part_vat] if c_part_vat else 0)
    ).apply(round_half_up)
    
    col_r = find_col_smart(df, ['CLAIM TYPE', 'CLAIMTYPE', '청구유형', '클레임유형', 'TYPE', '유형'], fallback_idx=17)
    col_v = find_col_smart(df, ['제목', 'TITLE', 'SUBJECT', '내용', '작업내용', '수리내용', 'DESCRIPTION', 'REMARK', '비고'], fallback_idx=21)

    excel_groups = defaultdict(list)
    for _, row in df.iterrows():
        claim_no = str(row.get(col_claim_no, '')).strip() if col_claim_no else ''
        if claim_no and claim_no != 'nan':
            r_val = str(row.get(col_r, '-')).strip() if col_r else '-'
            raw_v = str(row.get(col_v, '-')).strip() if col_v else '-'
            excel_groups[claim_no].append({
                'amount': int(row['Excel_Total']),
                'claim_type': r_val if r_val and r_val != 'nan' else '-',
                'v_desc': raw_v if raw_v and raw_v != 'nan' else '-'
            })
    return excel_groups

def load_pdf_mw(uploaded_file):
    pdf_groups = defaultdict(list)
    with pdfplumber.open(uploaded_file) as pdf:
        for page_num, page in enumerate(pdf.pages):
            if page_num % 2 != 0:
                continue
            text = page.extract_text()
            if not text:
                continue
            lines = text.split('\n')
            page_seen = set()
            for line in lines:
                line_stripped = line.strip()
                if line_stripped in page_seen:
                    continue
                page_seen.add(line_stripped)
                
                match = re.search(r'([A-Z]+\d+)', line_stripped)
                if match:
                    rep_order = match.group(1)
                    parts = line_stripped.split()
                    try:
                        total_str = parts[-1].replace(',', '.')
                        pdf_total_with_vat = round_half_up(float(total_str) * 1.1)
                        pdf_groups[rep_order].append(pdf_total_with_vat)
                    except ValueError:
                        continue
    return pdf_groups

def create_mw_excel_report(uploaded_file_mw, count, total_pdf, total_excel, total_diff):
    df_mw_raw = read_excel_smart_header(uploaded_file_mw)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "WARRANTY 수령내역"
    ws.print_title_rows = '1:3'
    
    target_headers = [
        "Claim No", "차량번호", "Job No", "완결일자", "청구일자",
        "공임청구액", "공임청구부가세", "부품청구액", "부품청구부가세",
        "공임입금액", "공임입금부가세", "부품입금액", "부품입금부가세"
    ]
    alias_dict = {
        "Claim No": ["CLAIM NO", "CLAIM", "클레임", "청구번호"],
        "차량번호": ["차량번호", "차량 번호", "CAR NO", "VEHICLE"],
        "Job No": ["JOB NO", "JOB", "작업번호"],
        "완결일자": ["완결일자", "완결일", "완결"],
        "청구일자": ["청구일자", "청구일"],
        "공임청구액": ["공임청구액", "공임청구", "공임 청구액"],
        "공임청구부가세": ["공임청구부가세", "공임청구 부가세", "공임부가세"],
        "부품청구액": ["부품청구액", "부품청구", "부품 청구액"],
        "부품청구부가세": ["부품청구부가세", "부품청구 부가세", "부품부가세"],
        "공임입금액": ["공임입금액", "공임입금", "공임승인액", "공임승인", "공임 입금액", "공임승인금액"],
        "공임입금부가세": ["공임입금부가세", "공임입금 부가세", "공임승인부가세"],
        "부품입금액": ["부품입금액", "부품입금", "부품승인액", "부품승인", "부품 입금액", "부품승인금액"],
        "부품입금부가세": ["부품입금부가세", "부품입금 부가세", "부품승인부가세"]
    }
    col_mapping = {th: find_col_smart(df_mw_raw, alias_dict.get(th, [th])) for th in target_headers}

    month_str = "6월"
    file_name = getattr(uploaded_file_mw, 'name', '')
    m_fn = re.search(r'20\d{2}(\d{2})', file_name)
    if m_fn:
        month_str = f"{int(m_fn.group(1))}월"
    else:
        for col in df_mw_raw.columns:
            if any(keyword in str(col).upper() for keyword in ['일자', 'DATE', '완결', '청구']):
                sample_dates = df_mw_raw[col].dropna().astype(str).tolist()
                for d in sample_dates:
                    m = re.search(r'-(\d{2})-', d) or re.search(r'/(\d{2})/', d)
                    if m:
                        month_str = f"{int(m.group(1))}월"
                        break
                if month_str != "6월":
                    break

    ws.merge_cells('A1:N1')
    ws['A1'] = f"{month_str} WARRANTY 수 령 내 역"
    ws['A1'].font = Font(size=22, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 40

    thin_border = Border(
        left=Side(style='thin', color='000000'), right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'), bottom=Side(style='thin', color='000000')
    )
    header_font = Font(size=10, bold=True)
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    ws.row_dimensions[3].height = 25
    cell_a = ws.cell(row=3, column=1, value="No.")
    cell_a.font = header_font
    cell_a.alignment = header_align
    cell_a.border = thin_border
    
    for col_pos, h_name in enumerate(target_headers, 2):
        cell = ws.cell(row=3, column=col_pos, value=h_name)
        cell.font = header_font
        cell.alignment = header_align
        cell.border = thin_border

    current_row = 4
    no_counter = 1
    for _, row in df_mw_raw.iterrows():
        if row.dropna().empty:
            continue
        ws.row_dimensions[current_row].height = 20
        c_no = ws.cell(row=current_row, column=1, value=no_counter)
        c_no.alignment = Alignment(horizontal='center', vertical='center')
        c_no.border = thin_border
        
        for col_pos, h_name in enumerate(target_headers, 2):
            cell = ws.cell(row=current_row, column=col_pos)
            mapped_col = col_mapping.get(h_name)
            if mapped_col and mapped_col in row and not pd.isna(row[mapped_col]):
                val = row[mapped_col]
                if isinstance(val, pd.Timestamp):
                    val = val.strftime('%Y-%m-%d')
                elif isinstance(val, str) and len(val) >= 10 and '00:00:00' in val:
                    val = val.split()[0]
                cell.value = val
                if isinstance(val, (int, float)):
                    cell.number_format = '#,##0'
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                else:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
            else:
                cell.value = ""
                cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
        current_row += 1
        no_counter += 1

    ws.row_dimensions[current_row].height = 25
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=2)
    c_sum = ws.cell(row=current_row, column=1, value="합계")
    c_sum.font = Font(bold=True)
    c_sum.alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=current_row, column=1).border = thin_border
    ws.cell(row=current_row, column=2).border = thin_border
    
    ws.merge_cells(start_row=current_row, start_column=3, end_row=current_row, end_column=4)
    c_cnt = ws.cell(row=current_row, column=3, value=f"댓수 : {count}")
    c_cnt.font = Font(bold=True)
    c_cnt.alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=current_row, column=3).border = thin_border
    ws.cell(row=current_row, column=4).border = thin_border
    
    ws.cell(row=current_row, column=5).border = thin_border
    ws.cell(row=current_row, column=6).border = thin_border
    
    c_g = ws.cell(row=current_row, column=7, value="총 실 수령액 :")
    c_g.font = Font(bold=True)
    c_g.alignment = Alignment(horizontal='center', vertical='center')
    c_g.border = thin_border
    
    c_h = ws.cell(row=current_row, column=8, value=total_pdf)
    c_h.font = Font(bold=True)
    c_h.number_format = '#,##0'
    c_h.alignment = Alignment(horizontal='right', vertical='center')
    c_h.border = thin_border
    
    c_i = ws.cell(row=current_row, column=9, value="총 청구 금액 :")
    c_i.font = Font(bold=True)
    c_i.alignment = Alignment(horizontal='center', vertical='center')
    c_i.border = thin_border
    
    c_j = ws.cell(row=current_row, column=10, value=total_excel)
    c_j.font = Font(bold=True)
    c_j.number_format = '#,##0'
    c_j.alignment = Alignment(horizontal='right', vertical='center')
    c_j.border = thin_border
    
    c_k = ws.cell(row=current_row, column=11, value="총 차액 :")
    c_k.font = Font(bold=True)
    c_k.alignment = Alignment(horizontal='center', vertical='center')
    c_k.border = thin_border
    
    c_l = ws.cell(row=current_row, column=12, value=total_diff)
    c_l.font = Font(bold=True)
    c_l.number_format = '#,##0'
    c_l.alignment = Alignment(horizontal='right', vertical='center')
    c_l.border = thin_border
    
    ws.merge_cells(start_row=current_row, start_column=13, end_row=current_row, end_column=14)
    c_mn = ws.cell(row=current_row, column=13, value="*부가세포함")
    c_mn.font = Font(bold=True)
    c_mn.alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=current_row, column=13).border = thin_border
    ws.cell(row=current_row, column=14).border = thin_border

    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row > 3 and cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output, month_str

# ────────────────────────────────────────────────────────
# 2️⃣ [모드 2] 쿠폰 보증 비교
# ────────────────────────────────────────────────────────
def load_excel_coupon_a(uploaded_file):
    df = read_excel_smart_header(uploaded_file)
    col_car = find_col_smart(df, ['차량번호', 'CAR NO', 'CAR_NO', 'VEHICLE'], fallback_idx=3)
    col_part = find_col_smart(df, ['부품청구', '부품'], fallback_idx=8)
    col_labour = find_col_smart(df, ['공임청구', '공임'], fallback_idx=9)
    if col_part: df[col_part] = pd.to_numeric(df[col_part], errors='coerce').fillna(0)
    if col_labour: df[col_labour] = pd.to_numeric(df[col_labour], errors='coerce').fillna(0)
    df['Calc_Total'] = (((df[col_part] if col_part else 0) + (df[col_labour] if col_labour else 0)) * 1.1).apply(round_half_up)
    
    col_r = find_col_smart(df, ['CLAIM TYPE', 'CLAIMTYPE', '청구유형', '클레임유형', 'TYPE', '유형'], fallback_idx=17)
    col_v = find_col_smart(df, ['제목', 'TITLE', 'SUBJECT', '내용', '작업내용', '수리내용', 'DESCRIPTION', 'REMARK', '비고'], fallback_idx=21)

    a_groups = defaultdict(list)
    for _, row in df.iterrows():
        car_no = str(row.get(col_car, '')).strip() if col_car else 'Unknown'
        if car_no and car_no != 'nan':
            r_val = str(row.get(col_r, '-')).strip() if col_r else '-'
            raw_v = str(row.get(col_v, '-')).strip() if col_v else '-'
            a_groups[car_no].append({
                'amount': int(row['Calc_Total']),
                'claim_type': r_val if r_val and r_val != 'nan' else '-',
                'v_desc': raw_v if raw_v and raw_v != 'nan' else '-'
            })
    return a_groups

def load_excel_coupon_b(uploaded_file):
    df = read_excel_smart_header(uploaded_file)
    col_car = find_col_smart(df, ['차량번호', 'CAR NO', 'CAR_NO', 'VEHICLE'], fallback_idx=6)
    col_total = find_col_smart(df, ['합계금액', '합계', 'TOTAL'], fallback_idx=18)
    if col_total: df[col_total] = pd.to_numeric(df[col_total], errors='coerce').fillna(0)
    
    col_r = find_col_smart(df, ['CLAIM TYPE', 'CLAIMTYPE', '청구유형', '클레임유형', 'TYPE', '유형'], fallback_idx=17)
    col_v = find_col_smart(df, ['제목', 'TITLE', 'SUBJECT', '내용', '작업내용', '수리내용', 'DESCRIPTION', 'REMARK', '비고'], fallback_idx=21)

    b_groups = defaultdict(list)
    for _, row in df.iterrows():
        car_no = str(row.get(col_car, '')).strip() if col_car else 'Unknown'
        if car_no and car_no != 'nan':
            r_val = str(row.get(col_r, '-')).strip() if col_r else '-'
            raw_v = str(row.get(col_v, '-')).strip() if col_v else '-'
            b_groups[car_no].append({
                'amount': round_half_up(row[col_total]) if col_total else 0,
                'claim_type': r_val if r_val and r_val != 'nan' else '-',
                'v_desc': raw_v if raw_v and raw_v != 'nan' else '-'
            })
    return b_groups

def create_coupon_excel_report(uploaded_file_a, uploaded_file_b, count, total_b, total_a, total_diff):
    df_a_raw = read_excel_smart_header(uploaded_file_a)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "쿠폰 청구 현황"
    ws.print_title_rows = '1:3'
    
    month_str = "8월"
    file_b_name = getattr(uploaded_file_b, 'name', '')
    m_fn = re.search(r'20\d{2}(\d{2})', file_b_name)
    if m_fn:
        month_str = f"{int(m_fn.group(1))}월"
    else:
        for col in df_a_raw.columns:
            if any(keyword in str(col) for keyword in ['일자', 'DATE', '승인', '청구', '입고', '출고']):
                sample_dates = df_a_raw[col].dropna().astype(str).tolist()
                for d in sample_dates:
                    m = re.search(r'-(\d{2})-', d) or re.search(r'/(\d{2})/', d)
                    if m:
                        month_str = f"{int(m.group(1))}월"
                        break
                if month_str != "8월":
                    break

    ws.merge_cells('A1:N1')
    ws['A1'] = f"{month_str} 쿠폰 청구 현황"
    ws['A1'].font = Font(size=22, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 40

    thin_border = Border(
        left=Side(style='thin', color='000000'), right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'), bottom=Side(style='thin', color='000000')
    )
    header_font = Font(size=10, bold=True)
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    headers = list(df_a_raw.columns)
    ws.row_dimensions[3].height = 25
    for col_idx, h_name in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx, value=str(h_name))
        cell.font = header_font
        cell.alignment = header_align
        cell.border = thin_border

    current_row = 4
    for _, row in df_a_raw.iterrows():
        ws.row_dimensions[current_row].height = 20
        for col_idx, val in enumerate(row, 1):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.value = "" if pd.isna(val) else val
            cell.border = thin_border
            if isinstance(val, (int, float)):
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal='right', vertical='center')
            else:
                cell.alignment = Alignment(horizontal='center', vertical='center')
        current_row += 1

    current_row += 2
    ws.row_dimensions[current_row].height = 30
    
    c_lbl1 = ws.cell(row=current_row, column=3, value="댓수 :")
    c_lbl1.font = Font(size=14, bold=True)
    c_lbl1.alignment = Alignment(horizontal='right', vertical='center')
    c_val1 = ws.cell(row=current_row, column=4, value=count)
    c_val1.font = Font(size=14, bold=True)
    c_val1.alignment = Alignment(horizontal='center', vertical='center')
    c_unit1 = ws.cell(row=current_row, column=5, value="대")
    c_unit1.font = Font(size=14, bold=True)
    c_unit1.alignment = Alignment(horizontal='left', vertical='center')

    c_lbl2 = ws.cell(row=current_row, column=7, value="총 청구 금액 :")
    c_lbl2.font = Font(size=14, bold=True)
    c_lbl2.alignment = Alignment(horizontal='right', vertical='center')
    c_val2 = ws.cell(row=current_row, column=9, value=total_b)
    c_val2.font = Font(size=14, bold=True)
    c_val2.number_format = '#,##0'
    c_val2.alignment = Alignment(horizontal='right', vertical='center')
    c_unit2 = ws.cell(row=current_row, column=10, value="원")
    c_unit2.font = Font(size=14, bold=True)
    c_unit2.alignment = Alignment(horizontal='left', vertical='center')
    c_vat1 = ws.cell(row=current_row, column=11, value="VAT 포함")
    c_vat1.font = Font(size=10, bold=True)
    c_vat1.alignment = Alignment(horizontal='left', vertical='center')

    current_row += 2
    ws.row_dimensions[current_row].height = 30
    
    c_lbl3 = ws.cell(row=current_row, column=3, value="차액 :")
    c_lbl3.font = Font(size=14, bold=True)
    c_lbl3.alignment = Alignment(horizontal='right', vertical='center')
    c_val3 = ws.cell(row=current_row, column=4, value=total_diff)
    c_val3.font = Font(size=14, bold=True)
    c_val3.number_format = '#,##0'
    c_val3.alignment = Alignment(horizontal='center', vertical='center')
    c_unit3 = ws.cell(row=current_row, column=5, value="원")
    c_unit3.font = Font(size=14, bold=True)
    c_unit3.alignment = Alignment(horizontal='left', vertical='center')

    c_lbl4 = ws.cell(row=current_row, column=7, value="총 입금 금액 :")
    c_lbl4.font = Font(size=14, bold=True)
    c_lbl4.alignment = Alignment(horizontal='right', vertical='center')
    c_val4 = ws.cell(row=current_row, column=9, value=total_a)
    c_val4.font = Font(size=14, bold=True)
    c_val4.number_format = '#,##0'
    c_val4.alignment = Alignment(horizontal='right', vertical='center')
    c_unit4 = ws.cell(row=current_row, column=10, value="원")
    c_unit4.font = Font(size=14, bold=True)
    c_unit4.alignment = Alignment(horizontal='left', vertical='center')
    c_vat2 = ws.cell(row=current_row, column=11, value="VAT 포함")
    c_vat2.font = Font(size=10, bold=True)
    c_vat2.alignment = Alignment(horizontal='left', vertical='center')

    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row > 3 and cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_len + 5, 14)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output, month_str

# ────────────────────────────────────────────────────────
# 3️⃣ [모드 3] 공임코드 비교
# ────────────────────────────────────────────────────────
def parse_labor_lines(text):
    code_map = defaultdict(list)
    if not text:
        return code_map
    pattern = re.compile(r'([A-Za-z0-9]{3}-[A-Za-z0-9]{2}-[A-Za-z0-9]{1,4})')
    for line in text.split('\n'):
        line_clean = line.strip()
        if not line_clean:
            continue
        match = pattern.search(line_clean)
        if match:
            code = match.group(1).upper()
            code_map[code].append(line_clean)
    return code_map

# ────────────────────────────────────────────────────────
# 🖥️ 본문 화면 렌더링
# ────────────────────────────────────────────────────────
if mode == "MW 보증 비교":
    st.markdown("### 🔍 PDF(홀수페이지)와 엑셀의 금액을 각각 계산 후 반올림 처리하여 순차 정렬 대조합니다.")
    st.write("")
    
    col1, col2 = st.columns(2)
    with col1:
        st.markdown("### 1. PDF 파일을 선택하세요 (예시 : DEALER_CREDITNOTE_6755)")
        pdf_file = st.file_uploader("PDF 파일 업로드", type=["pdf"], key="mw_pdf", label_visibility="collapsed")
    with col2:
        st.markdown("### 2. 엑셀 파일을 선택하세요 (예시 : 보증청구현황 [ 항목 조정 가능 ]_2026)")
        excel_file = st.file_uploader("엑셀 파일 업로드", type=["xlsx"], key="mw_excel", label_visibility="collapsed")
        
    if pdf_file and excel_file:
        with st.spinner("MW 보증 데이터 교차 대조 중..."):
            excel_groups = load_excel_mw(excel_file)
            pdf_groups = load_pdf_mw(pdf_file)
            
            matched_results = []
            diff_over_100_results = []
            
            all_orders = sorted(list(set(list(excel_groups.keys()) + list(pdf_groups.keys()))))
            total_pdf_sum = 0
            total_excel_sum = 0
            total_diff_100_sum = 0
            
            for order in all_orders:
                p_amts = pdf_groups.get(order, [])
                e_items = excel_groups.get(order, [])
                max_len = max(len(p_amts), len(e_items))
                
                for i in range(max_len):
                    p_amt = p_amts[i] if i < len(p_amts) else None
                    e_item = e_items[i] if i < len(e_items) else None
                    e_amt = e_item['amount'] if e_item else None
                    r_val = e_item['claim_type'] if e_item else '-'
                    v_val = e_item['v_desc'] if e_item else '-'
                    
                    order_label = f"{order} ({i+1})" if max_len > 1 else order
                    
                    if p_amt is not None: total_pdf_sum += p_amt
                    if e_amt is not None: total_excel_sum += e_amt
                    
                    diff_val = 0
                    if p_amt is not None and e_amt is not None:
                        diff_val = p_amt - e_amt
                        row_dict = {
                            '주문번호': order_label,
                            'PDF 금액 (실 수령액)': f"{p_amt:,}원",
                            'DMS 금액 (청구 금액)': f"{e_amt:,}원",
                            '차액': f"{diff_val:,}원" if diff_val != 0 else "0원"
                        }
                    elif p_amt is not None:
                        diff_val = p_amt
                        row_dict = {
                            '주문번호': order_label,
                            'PDF 금액 (실 수령액)': f"{p_amt:,}원",
                            'DMS 금액 (청구 금액)': "-",
                            '차액': f"{diff_val:,}원"
                        }
                    elif e_amt is not None:
                        diff_val = -e_amt
                        row_dict = {
                            '주문번호': order_label,
                            'PDF 금액 (실 수령액)': "-",
                            'DMS 금액 (청구 금액)': f"{e_amt:,}원",
                            '차액': f"{diff_val:,}원"
                        }
                    matched_results.append(row_dict)
                    
                    if abs(diff_val) >= 100:
                        total_diff_100_sum += diff_val
                        diff_over_100_results.append({
                            '주문번호': order_label,
                            'Claim Type': r_val,
                            '제목': v_val,
                            '차액': f"{diff_val:,}원"
                        })
            
            total_diff_sum = total_pdf_sum - total_excel_sum
            matched_results.append({
                '주문번호': "★ 총합계",
                'PDF 금액 (실 수령액)': f"{total_pdf_sum:,}원",
                'DMS 금액 (청구 금액)': f"{total_excel_sum:,}원",
                '차액': f"{total_diff_sum:,}원"
            })
            
            res_df = pd.DataFrame(matched_results)
            res_df.index = [str(i) for i in range(1, len(res_df))] + [""]
            
            st.subheader("📌 분석 요약 결과")
            m_col1, m_col2, m_col3, m_col4 = st.columns(4)
            mw_count = len(res_df) - 1
            m_col1.metric("총 대조 건수", f"{mw_count} 건")
            m_col2.metric("PDF 총 합계 금액", f"{total_pdf_sum:,}원")
            m_col3.metric("DMS 총 합계 금액", f"{total_excel_sum:,}원")
            m_col4.metric("최종 총 차이 금액", f"{total_diff_sum:,}원", delta=f"{total_diff_sum:,}원" if total_diff_sum != 0 else None)
            
            mw_excel_data, mw_month_name = create_mw_excel_report(
                excel_file, mw_count, total_pdf_sum, total_excel_sum, total_diff_sum
            )
            
            st.write("")
            st.download_button(
                label=f"📥 [{mw_month_name} WARRANTY 수령내역] 엑셀 보고서 다운로드",
                data=mw_excel_data,
                file_name=f"{mw_month_name}_WARRANTY_수령내역_보고서.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )
            
            st.write("")
            if diff_over_100_results:
                diff_list_with_total = list(diff_over_100_results)
                diff_list_with_total.append({
                    '주문번호': "★ 총합계",
                    'Claim Type': "-",
                    '제목': "-",
                    '차액': f"{total_diff_100_sum:,}원"
                })
                diff_df = pd.DataFrame(diff_list_with_total)
                diff_df.index = [str(i) for i in range(1, len(diff_df))] + [""]
            else:
                diff_df = pd.DataFrame(columns=['주문번호', 'Claim Type', '제목', '차액'])

            render_side_by_side_tables(res_df, diff_df)

elif mode == "쿠폰 보증 비교":
    st.markdown("### 🔍 공지된 쿠폰 금액 과 DMS 에서 출력된 쿠폰 금액을 정밀 매칭합니다. (차량번호 기준)")
    st.write("")
    
    col1, col2 = st.columns(2)
    with col1:
        st.markdown("### 1. 공지된 쿠폰 파일 (예시 : IR_JJ_Aug)")
        file_a = st.file_uploader("공지된 쿠폰 파일 업로드", type=["xlsx"], key="cp_a", label_visibility="collapsed")
    with col2:
        st.markdown("### 2. DMS 쿠폰파일 (예시 : 쿠폰청구관리_20260818085441)")
        file_b = st.file_uploader("DMS 쿠폰파일 업로드", type=["xlsx"], key="cp_b", label_visibility="collapsed")
        
    if file_a and file_b:
        with st.spinner("쿠폰 보증 엑셀 간 교차 대조 중..."):
            a_groups = load_excel_coupon_a(file_a)
            b_groups = load_excel_coupon_b(file_b)
            
            matched_results = []
            diff_over_100_results = []
            
            all_cars = sorted(list(set(list(a_groups.keys()) + list(b_groups.keys()))))
            total_a_sum = 0
            total_b_sum = 0
            total_diff_100_sum = 0
            
            for car in all_cars:
                a_items = a_groups.get(car, [])
                b_items = b_groups.get(car, [])
                max_len = max(len(a_items), len(b_items))
                
                for i in range(max_len):
                    a_item = a_items[i] if i < len(a_items) else None
                    b_item = b_items[i] if i < len(b_items) else None
                    a_amt = a_item['amount'] if a_item else None
                    b_amt = b_item['amount'] if b_item else None
                    
                    r_val = b_item['claim_type'] if b_item else (a_item['claim_type'] if a_item else '-')
                    v_val = b_item['v_desc'] if b_item else (a_item['v_desc'] if a_item else '-')
                    
                    car_label = f"{car} ({i+1})" if max_len > 1 else car
                    
                    if a_amt is not None: total_a_sum += a_amt
                    if b_amt is not None: total_b_sum += b_amt
                    
                    diff_val = 0
                    if a_amt is not None and b_amt is not None:
                        diff_val = a_amt - b_amt
                        row_dict = {
                            '차량번호': car_label,
                            '공지된 쿠폰 금액 ( 입금 금액 )': f"{a_amt:,}원",
                            'DMS 쿠폰파일 ( 청구 금액 ) ': f"{b_amt:,}원",
                            '차액': f"{diff_val:,}원" if diff_val != 0 else "0원"
                        }
                    elif a_amt is not None:
                        diff_val = a_amt
                        row_dict = {
                            '차량번호': car_label,
                            '공지된 쿠폰 금액 ( 입금 금액 )': f"{a_amt:,}원",
                            'DMS 쿠폰파일 ( 청구 금액 ) ': "-",
                            '차액': f"{diff_val:,}원"
                        }
                    elif b_amt is not None:
                        diff_val = -b_amt
                        row_dict = {
                            '차량번호': car_label,
                            '공지된 쿠폰 금액 ( 입금 금액 )': "-",
                            'DMS 쿠폰파일 ( 청구 금액 ) ': f"{b_amt:,}원",
                            '차액': f"{diff_val:,}원"
                        }
                    matched_results.append(row_dict)
                    
                    if abs(diff_val) >= 100:
                        total_diff_100_sum += diff_val
                        diff_over_100_results.append({
                            '차량번호': car_label,
                            'Claim Type': r_val,
                            '제목': v_val,
                            '차액': f"{diff_val:,}원"
                        })
            
            total_diff_sum = total_a_sum - total_b_sum
            matched_results.append({
                '차량번호': "★ 총합계",
                '공지된 쿠폰 금액 ( 입금 금액 )': f"{total_a_sum:,}원",
                'DMS 쿠폰파일 ( 청구 금액 ) ': f"{total_b_sum:,}원",
                '차액': f"{total_diff_sum:,}원"
            })
            
            res_df = pd.DataFrame(matched_results)
            res_df.index = [str(i) for i in range(1, len(res_df))] + [""]
            
            st.subheader("📌 분석 요약 결과")
            m_col1, m_col2, m_col3, m_col4 = st.columns(4)
            total_count = len(res_df) - 1
            m_col1.metric("총 대조 건수", f"{total_count} 건")
            m_col2.metric("공지 쿠폰 총 합계", f"{total_a_sum:,}원")
            m_col3.metric("DMS 쿠폰 총 합계", f"{total_b_sum:,}원")
            m_col4.metric("최종 총 차이 금액", f"{total_diff_sum:,}원", delta=f"{total_diff_sum:,}원" if total_diff_sum != 0 else None)
            
            excel_data, month_name = create_coupon_excel_report(
                file_a, file_b, total_count, total_b_sum, total_a_sum, total_diff_sum
            )
            
            st.write("")
            st.download_button(
                label=f"📥 [{month_name} 쿠폰 청구 현황] 엑셀 보고서 다운로드",
                data=excel_data,
                file_name=f"{month_name}_쿠폰_청구_현황_보고서.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )
            
            st.write("")
            if diff_over_100_results:
                diff_list_with_total = list(diff_over_100_results)
                diff_list_with_total.append({
                    '차량번호': "★ 총합계",
                    'Claim Type': "-",
                    '제목': "-",
                    '차액': f"{total_diff_100_sum:,}원"
                })
                diff_df = pd.DataFrame(diff_list_with_total)
                diff_df.index = [str(i) for i in range(1, len(diff_df))] + [""]
            else:
                diff_df = pd.DataFrame(columns=['차량번호', 'Claim Type', '제목', '차액'])

            render_side_by_side_tables(res_df, diff_df)

else:
    st.markdown("### 🔍 A 그룹과 B 그룹에 복사한 공임 텍스트를 붙여넣은 뒤, **[비교진행]** 버튼을 누르면 `3자리-2자리-1~4자리` 형태의 공임코드 중복을 찾아냅니다.")
    st.write("")
    
    col_a, col_b = st.columns(2)
    with col_a:
        st.markdown("### A 그룹 내용 붙여넣기")
        text_a = st.text_area("A그룹", height=280, placeholder="예시:\n900-00-B   Engine hood open and close   1   7", label_visibility="collapsed")
    with col_b:
        st.markdown("### B 그룹 내용 붙여넣기")
        text_b = st.text_area("B그룹", height=280, placeholder="예시:\n211-13-G11   Cover over engine remove-install   1   12", label_visibility="collapsed")
        
    start_compare = st.button("🔍 비교진행", use_container_width=True, type="primary")
    
    if start_compare:
        if not text_a.strip() and not text_b.strip():
            st.warning("⚠️ A그룹 또는 B그룹에 내용을 먼저 붙여넣어 주세요.")
        else:
            map_a = parse_labor_lines(text_a)
            map_b = parse_labor_lines(text_b)
            
            duplicate_codes = sorted(list(set(map_a.keys()) & set(map_b.keys())))
            only_a = sorted(list(set(map_a.keys()) - set(map_b.keys())))
            only_b = sorted(list(set(map_b.keys()) - set(map_a.keys())))
            
            st.divider()
            st.subheader("📌 비교 요약 결과")
            c1, c2, c3 = st.columns(3)
            c1.metric("중복된 공임코드", f"{len(duplicate_codes)} 건", delta="중복 발견" if duplicate_codes else None)
            c2.metric("A그룹 고유 항목", f"{len(only_a)} 건")
            c3.metric("B그룹 고유 항목", f"{len(only_b)} 건")
            
            if duplicate_codes:
                st.markdown("### 🚨 중복 발견 내역")
                dup_rows = []
                for idx, code in enumerate(duplicate_codes, 1):
                    lines_a_str = " | ".join(map_a[code])
                    lines_b_str = " | ".join(map_b[code])
                    dup_rows.append({
                        "중복 공임코드": code,
                        "A그룹 원본 내용": lines_a_str,
                        "B그룹 원본 내용": lines_b_str
                    })
                df_dup = pd.DataFrame(dup_rows)
                df_dup.index = range(1, len(df_dup) + 1)
                
                main_headers = ["No."] + list(df_dup.columns)
                main_tbody = []
                for idx, row in df_dup.iterrows():
                    main_tbody.append(f'<tr><td class="col-no">{idx}</td><td class="col-id copyable" onclick="copyCell(this)">{row.iloc[0]}</td><td class="col-amt">{row.iloc[1]}</td><td class="col-amt">{row.iloc[2]}</td></tr>')
                
                css_dup = """
                  * { box-sizing: border-box; font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, sans-serif; }
                  body { background-color: transparent; color: #f8fafc; }
                  table { border-collapse: collapse; width: 100%; font-size: 15px; user-select: text; }
                  th { background-color: #1e293b; color: #fff; padding: 10px 12px; border: 1px solid #334155; text-align: center; }
                  td { padding: 8px 12px; border: 1px solid #334155; }
                  td.copyable { cursor: pointer; }
                  td.copyable:hover { background-color: rgba(14, 165, 233, 0.2) !important; }
                  #toast { visibility: hidden; position: fixed; top: 10px; left: 50%; transform: translateX(-50%); background-color: #0284c7; color: #fff; padding: 8px 16px; border-radius: 6px; font-weight: bold; z-index: 99999; }
                  #toast.show { visibility: visible; }
                """
                
                js_dup = """
                  function copyCell(el) {
                    const text = el.innerText.trim();
                    if (!text || text === '-') return;
                    navigator.clipboard.writeText(text);
                    const toast = document.getElementById('toast');
                    toast.innerText = '📋 복사 완료: ' + text;
                    toast.className = 'show';
                    setTimeout(function() { toast.className = ''; }, 1200);
                  }
                """
                
                table_html = (
                    '<!DOCTYPE html><html><head><meta charset="utf-8" />'
                    f'<style>{css_dup}</style></head><body>'
                    '<div id="toast">📋 복사 완료!</div>'
                    '<table><thead><tr>'
                    f'<th>{main_headers[0]}</th><th>{main_headers[1]}</th><th>{main_headers[2]}</th><th>{main_headers[3]}</th>'
                    '</tr></thead>'
                    f'<tbody>{"".join(main_tbody)}</tbody></table>'
                    f'<script>{js_dup}</script>'
                    '</body></html>'
                )
                components.html(table_html, height=min(600, len(df_dup) * 44 + 80), scrolling=True)
            else:
                st.success("✅ A그룹과 B그룹 간에 중복된 공임코드가 없습니다.")
